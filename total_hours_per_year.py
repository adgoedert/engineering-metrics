"""
Annual WBSO hours: sum the monthly consolidated workbooks for a year.

Reads consolidated-wbso-hours-<year>-01.xlsx .. -12.xlsx, takes the all-teams
total from each Summary sheet, and prints the yearly total.

Getting that number needs care. openpyxl writes formulas with no cached value,
so the "ALL TEAMS" cell reads back as None until someone opens the file in Excel
and saves it. Summing those Nones naively yields 0 — a plausible-looking wrong
answer. So each file is read in this order:

  1. the cached value of the ALL TEAMS row, if Excel has populated it
  2. the cached per-team cells in the same column, summed
  3. the Detail sheet's literal hours, which is what those formulas aggregate

Route 3 always works, because Detail holds numbers rather than formulas. Where
both a cached total and the Detail figure are available they are cross-checked
and any disagreement is reported rather than silently preferred.
"""

import argparse
import os

from openpyxl import load_workbook
from tabulate import tabulate

DEFAULT_YEAR = 2025
FILE_PATTERN = "consolidated-wbso-hours-{year}-{month:02d}.xlsx"
MONTH_NAMES = ["January", "February", "March", "April", "May", "June", "July",
               "August", "September", "October", "November", "December"]

SUMMARY_SHEET = "Summary"
DETAIL_SHEET = "Detail"
HOURS_HEADER = "WBSO hours"
TOTAL_LABEL = "ALL TEAMS"
DETAIL_TEAM_COL = 1
DETAIL_HOURS_COL = 7


def find_hours_column(ws):
    """Locate the 'WBSO hours' header, returning (row, column). None if absent."""
    for row in ws.iter_rows(min_row=1, max_row=12):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip().lower() == HOURS_HEADER.lower():
                return cell.row, cell.column
    return None, None


def find_total_row(ws, header_row):
    for r in range(header_row + 1, ws.max_row + 1):
        value = ws.cell(row=r, column=1).value
        if isinstance(value, str) and value.strip().upper() == TOTAL_LABEL:
            return r
    return None


def detail_total(ws_detail):
    """Sum the literal hours on Detail, with the per-team split."""
    total = 0.0
    per_team = {}
    rows = 0
    for r in range(2, ws_detail.max_row + 1):
        team = ws_detail.cell(row=r, column=DETAIL_TEAM_COL).value
        hours = ws_detail.cell(row=r, column=DETAIL_HOURS_COL).value
        if team and isinstance(hours, (int, float)):
            total += float(hours)
            per_team[team] = per_team.get(team, 0.0) + float(hours)
            rows += 1
    return total, per_team, rows


def read_month(path):
    """Extract one month's all-teams total. Never raises; problems are reported."""
    out = {
        "path": path, "hours": None, "source": "", "teams": 0,
        "detail_rows": 0, "per_team": {}, "note": "", "error": "",
    }

    if not os.path.exists(path):
        out["error"] = "file not found"
        return out

    try:
        wb_values = load_workbook(path, data_only=True)
    except Exception as exc:
        out["error"] = f"cannot open ({type(exc).__name__}: {exc})"
        return out

    if SUMMARY_SHEET not in wb_values.sheetnames:
        out["error"] = f"no '{SUMMARY_SHEET}' sheet (found: {', '.join(wb_values.sheetnames)})"
        return out

    summary = wb_values[SUMMARY_SHEET]
    header_row, hours_col = find_hours_column(summary)
    if hours_col is None:
        out["error"] = f"no '{HOURS_HEADER}' column on {SUMMARY_SHEET}"
        return out

    total_row = find_total_row(summary, header_row)

    # Team rows sit between the header and the total row.
    team_rows = []
    last = (total_row - 1) if total_row else summary.max_row
    for r in range(header_row + 1, last + 1):
        name = summary.cell(row=r, column=1).value
        if isinstance(name, str) and name.strip() and name.strip().upper() != TOTAL_LABEL:
            team_rows.append((r, name.strip()))
    out["teams"] = len(team_rows)

    derived = None
    if DETAIL_SHEET in wb_values.sheetnames:
        derived, per_team, n_rows = detail_total(wb_values[DETAIL_SHEET])
        out["per_team"] = per_team
        out["detail_rows"] = n_rows
        # A Detail team missing from Summary would make the two disagree.
        orphans = set(per_team) - {name for _, name in team_rows}
        if orphans:
            out["note"] = f"Detail teams absent from Summary: {', '.join(sorted(orphans))}"

    # 1. cached ALL TEAMS cell
    cached_total = None
    if total_row:
        value = summary.cell(row=total_row, column=hours_col).value
        if isinstance(value, (int, float)):
            cached_total = float(value)

    # 2. cached per-team cells
    per_row_total = None
    if cached_total is None and team_rows:
        values = [summary.cell(row=r, column=hours_col).value for r, _ in team_rows]
        if values and all(isinstance(v, (int, float)) for v in values):
            per_row_total = float(sum(values))

    if cached_total is not None:
        out["hours"] = cached_total
        out["source"] = "Summary total (cached)"
        if derived is not None and abs(derived - cached_total) > 0.05:
            out["note"] = (f"cached total {cached_total:.1f} disagrees with Detail "
                           f"{derived:.1f} — workbook may be stale").strip()
    elif per_row_total is not None:
        out["hours"] = per_row_total
        out["source"] = "Summary team rows (cached)"
        if derived is not None and abs(derived - per_row_total) > 0.05:
            out["note"] = (f"team rows {per_row_total:.1f} disagree with Detail "
                           f"{derived:.1f} — workbook may be stale").strip()
    elif derived is not None:
        out["hours"] = derived
        out["source"] = "Detail (derived — formulas uncached)"
    else:
        out["error"] = "no usable values: formulas uncached and no Detail sheet"

    return out


def main():
    parser = argparse.ArgumentParser(
        description="Sum WBSO hours across the monthly consolidated workbooks for a year."
    )
    parser.add_argument("--year", type=int, default=DEFAULT_YEAR,
                        help=f"Year to total (default: {DEFAULT_YEAR})")
    parser.add_argument("--dir", default=".", help="Directory holding the workbooks (default: .)")
    parser.add_argument("--per-team", action="store_true",
                        help="Also break the annual total down per team")
    args = parser.parse_args()

    print(f"\n=== Total WBSO hours per year — {args.year} ===\n")

    results = []
    for month in range(1, 13):
        path = os.path.join(args.dir, FILE_PATTERN.format(year=args.year, month=month))
        res = read_month(path)
        res["month"] = month
        results.append(res)
        print(f"  [{month:02d}] {os.path.basename(path):40} "
              f"{'ERROR: ' + res['error'] if res['error'] else f'{res['hours']:>9.1f} h'}")

    print()
    table = []
    for res in results:
        table.append([
            f"{res['month']:02d}",
            MONTH_NAMES[res["month"] - 1],
            res["teams"] or "-",
            res["detail_rows"] or "-",
            f"{res['hours']:.1f}" if res["hours"] is not None else "-",
            res["source"] or res["error"],
        ])
    print(tabulate(table,
                   headers=["MM", "Month", "Teams", "Detail rows", "WBSO hours", "Source"],
                   tablefmt="rounded_outline", floatfmt=".1f"))

    found = [r for r in results if r["hours"] is not None]
    missing = [r for r in results if r["hours"] is None]
    total = sum(r["hours"] for r in found)

    print(f"\nMonths read      : {len(found)}/12")
    if missing:
        for r in missing:
            print(f"  ! {MONTH_NAMES[r['month'] - 1]}: {r['error']}")
    notes = [r for r in found if r["note"]]
    if notes:
        print()
        for r in notes:
            print(f"  ! {MONTH_NAMES[r['month'] - 1]}: {r['note']}")

    derived_only = [r for r in found if r["source"].startswith("Detail")]
    if derived_only:
        print(f"\nNote: {len(derived_only)} of {len(found)} month(s) had uncached formulas, so the")
        print("total came from the Detail sheet's literal hours. That is the same figure")
        print("the Summary formulas compute — open a file in Excel to see them agree.")

    print(f"\n{'=' * 58}")
    print(f"  TOTAL WBSO HOURS, ALL TEAMS, {args.year}: {total:>14,.1f} h")
    print(f"  Equivalent 8h days                    : {total / 8:>14,.1f} d")
    print(f"{'=' * 58}")

    if args.per_team:
        annual = {}
        for res in found:
            for team, hours in res["per_team"].items():
                annual[team] = annual.get(team, 0.0) + hours
        if annual:
            rows = [[t, f"{h:.1f}", f"{h / 8:.1f}", f"{h / total * 100:.1f}%" if total else "-"]
                    for t, h in sorted(annual.items(), key=lambda x: -x[1])]
            rows.append(["TOTAL", f"{sum(annual.values()):.1f}",
                         f"{sum(annual.values()) / 8:.1f}", "100.0%"])
            print(f"\n--- Per team, {args.year} ---")
            print(tabulate(rows, headers=["Team", "Hours", "Days", "Share"],
                           tablefmt="rounded_outline", floatfmt=".1f"))


if __name__ == "__main__":
    main()
