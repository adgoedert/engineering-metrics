"""
Writing the consolidated WBSO hours report to a .xlsx workbook.

Layout follows the usual spreadsheet discipline: one flat sheet of raw
per-ticket hours as the single source of truth, and every aggregate above it
expressed as a formula referencing that sheet. Change a number on Detail and
the epic subtotals, engineer totals and WBSO shares all follow.

  Assumptions   parameters the run used, incl. the hours-per-day constant that
                every capacity formula references
  Summary       one row per team: hours, capacity, WBSO share, status
  <Team name>   one sheet per team — hours by engineer and epic, then a
                per-engineer block with capacity and share
  Detail        one row per team / engineer / epic / ticket — the input data
  Diagnostics   epics that need attention in the mapping CSV

Sheet order puts Summary before the team sheets so the workbook opens on
something readable, and Detail after them since it is the raw feed.
"""

import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=14)
SECTION_FONT = Font(name=FONT, bold=True, size=11)
BODY_FONT = Font(name=FONT, size=10)
BOLD_FONT = Font(name=FONT, bold=True, size=10)
INPUT_FONT = Font(name=FONT, size=10, color="0000FF")  # blue = hardcoded input
NOTE_FONT = Font(name=FONT, size=9, italic=True, color="666666")
ASSUMPTION_FILL = PatternFill("solid", fgColor="FFFF00")
SUBTOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HOURS_FMT = "0.0"
PCT_FMT = "0.0%"

# Assumptions!B5 holds hours-per-day; every capacity formula references it.
HOURS_PER_DAY_REF = "Assumptions!$B$5"

# Detail sheet layout. Data starts at DETAIL_FIRST_ROW; columns are fixed so
# formulas on other sheets can be built before the sheet itself is written.
DETAIL_FIRST_ROW = 4
COL_TEAM, COL_ENG, COL_EPIC_KEY, COL_EPIC, COL_TICKET, COL_SUMMARY, COL_HOURS = range(1, 8)

INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")


def safe_sheet_name(name: str, used: set) -> str:
    """Excel caps sheet names at 31 chars, forbids []:*?/\\ and duplicates."""
    cleaned = INVALID_SHEET_CHARS.sub("-", name).strip() or "Team"
    cleaned = cleaned[:31]
    candidate, n = cleaned, 2
    while candidate.lower() in used:
        suffix = f" ({n})"
        candidate = cleaned[: 31 - len(suffix)] + suffix
        n += 1
    used.add(candidate.lower())
    return candidate


def _write_header(ws, row, headers, widths=None, freeze=True):
    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX
    if widths:
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    if freeze:
        ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _flatten(results):
    """
    Per-team results into flat Detail rows, sorted. Sorting happens in Python:
    LibreOffice cannot evaluate SORT().
    """
    rows = []
    for res in results:
        totals = res.get("totals") or {}
        summaries = res.get("summaries") or {}
        epic_of = (res.get("diagnostics") or {}).get("epic_of", {})
        for engineer, epic_map in totals.items():
            for epic_label, tickets in epic_map.items():
                for ticket, hours in tickets.items():
                    epic_key = (epic_of.get(ticket) or (None, epic_label))[0] or ""
                    rows.append({
                        "team": res["team"],
                        "engineer": engineer,
                        "epic_key": epic_key,
                        "epic_label": epic_label,
                        "ticket": ticket,
                        "summary": summaries.get(ticket, ""),
                        "hours": hours,
                    })

    eng_totals = {}
    for r in rows:
        key = (r["team"], r["engineer"])
        eng_totals[key] = eng_totals.get(key, 0) + r["hours"]
    rows.sort(key=lambda r: (r["team"], -eng_totals[(r["team"], r["engineer"])],
                             r["engineer"], r["epic_key"], -r["hours"], r["ticket"]))
    return rows


def _detail_range(col, n_rows):
    """Absolute Detail range for a column, e.g. Detail!$G$4:$G$27."""
    letter = get_column_letter(col)
    last = DETAIL_FIRST_ROW + max(n_rows, 1) - 1
    return f"Detail!${letter}${DETAIL_FIRST_ROW}:${letter}${last}"


# -- sheets ----------------------------------------------------------------


def _sheet_assumptions(wb, meta):
    ws = wb.create_sheet("Assumptions")
    ws["A1"] = "Consolidated WBSO Hours — run parameters"
    ws["A1"].font = TITLE_FONT
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 52

    ws["A3"] = "Parameter"
    ws["B3"] = "Value"
    for c in ("A3", "B3"):
        ws[c].font = HEADER_FONT
        ws[c].fill = HEADER_FILL
        ws[c].border = BOX

    # Row 5 is referenced by every capacity formula — keep it at this address.
    entries = [
        ("Hours per engineer per day", meta["hours_per_day"]),
        ("Reporting month", meta["month"]),
        ("Teams in scope", meta["team_count"]),
        ("Teams with WBSO hours", meta["teams_with_data"]),
        ("Epic filter", meta["filter_scope"]),
        ("Mapping CSV", meta["mapping_file"]),
        ("Generated by", "engineering-metrics / consolidated-wbso-hours"),
    ]
    row = 5
    for label, value in entries:
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = INPUT_FONT
        if row == 5:
            cell.fill = ASSUMPTION_FILL
            cell.number_format = HOURS_FMT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Method").font = SECTION_FONT
    notes = [
        "A weekday on which an engineer held a ticket while it was In Progress "
        "makes that ticket active for them that day.",
        "Each engineer-day is worth at most the hours-per-day figure above, split "
        "equally across every ticket active that day.",
        "The split denominator is ALL tickets held that day, including non-WBSO "
        "ones. Filtering to WBSO removes rows but never inflates the hours "
        "remaining — important for a subsidy claim.",
        "An engineer working for two teams in the month is counted under each, so "
        "team capacity figures do not sum to a person-month.",
        "Weekends are excluded. Public holidays are NOT excluded.",
        "Hours on the Detail sheet are computed by the tool from the Jira "
        "changelog; they are inputs here, not sheet-derived. Every total above "
        "them is a formula.",
    ]
    for note in notes:
        row += 1
        cell = ws.cell(row=row, column=1, value="• " + note)
        cell.font = NOTE_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.row_dimensions[row].height = 28
    return ws


def _sheet_summary(wb, results, n_rows, sheet_names):
    ws = wb.create_sheet("Summary")
    ws["A1"] = "WBSO hours per team"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Hours are SUMIFS over Detail. Capacity = active engineer-days x "
                "hours-per-day (Assumptions!B5).")
    ws["A2"].font = NOTE_FONT
    _write_header(
        ws, 4,
        ["Team", "Jira key", "Sheet", "Engineers", "WBSO epics",
         "WBSO hours", "Active eng-days", "Capacity hours", "WBSO share", "Status"],
        [26, 10, 26, 11, 11, 13, 16, 15, 12, 34],
    )

    hours_col = _detail_range(COL_HOURS, n_rows)
    team_col = _detail_range(COL_TEAM, n_rows)

    row = 5
    first = row
    for res in results:
        team = res["team"]
        has_data = bool(res.get("totals"))
        ws.cell(row=row, column=1, value=team).font = BODY_FONT
        ws.cell(row=row, column=2, value=res.get("project", "")).font = BODY_FONT
        ws.cell(row=row, column=3, value=sheet_names.get(team, "")).font = BODY_FONT

        day_counts = res.get("day_counts") or {}
        ws.cell(row=row, column=4, value=len(day_counts)).font = INPUT_FONT
        ws.cell(row=row, column=5, value=res.get("wbso_epic_count", 0)).font = INPUT_FONT

        hours = ws.cell(row=row, column=6)
        hours.value = f"=SUMIFS({hours_col},{team_col},$A{row})" if has_data else 0
        hours.number_format = HOURS_FMT
        hours.font = BODY_FONT

        eng_days = ws.cell(row=row, column=7, value=sum(day_counts.values()))
        eng_days.font = INPUT_FONT

        capacity = ws.cell(row=row, column=8, value=f"=$G{row}*{HOURS_PER_DAY_REF}")
        capacity.number_format = HOURS_FMT
        capacity.font = BODY_FONT

        share = ws.cell(row=row, column=9, value=f"=IFERROR($F{row}/$H{row},0)")
        share.number_format = PCT_FMT
        share.font = BODY_FONT

        ws.cell(row=row, column=10, value=res.get("status", "")).font = (
            BODY_FONT if has_data else NOTE_FONT
        )
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = BOX
        row += 1

    total_row = row + 1
    ws.cell(row=total_row, column=1, value="ALL TEAMS").font = BOLD_FONT
    for col in (4, 5, 6, 7, 8):
        letter = get_column_letter(col)
        cell = ws.cell(row=total_row, column=col, value=f"=SUM({letter}{first}:{letter}{row - 1})")
        cell.font = BOLD_FONT
        cell.number_format = HOURS_FMT if col in (6, 8) else "0"
        cell.fill = SUBTOTAL_FILL
    share = ws.cell(row=total_row, column=9, value=f"=IFERROR($F{total_row}/$H{total_row},0)")
    share.font = BOLD_FONT
    share.number_format = PCT_FMT
    share.fill = SUBTOTAL_FILL
    ws.auto_filter.ref = f"A4:J{row - 1}"
    return ws


def _sheet_team(wb, res, rows, n_rows, sheet_name):
    """One team: hours by engineer and epic, then a per-engineer block."""
    ws = wb.create_sheet(sheet_name)
    team = res["team"]
    ws["A1"] = f"{team} — WBSO hours, {res['month']}"
    ws["A1"].font = TITLE_FONT

    ws["A2"] = team
    ws["A2"].font = INPUT_FONT
    ws["C2"] = "^ SUMIFS key for this sheet; must match the Team column on Detail."
    ws["C2"].font = NOTE_FONT

    if not res.get("totals"):
        ws["A4"] = res.get("status") or "No WBSO hours for this team in the reporting month."
        ws["A4"].font = BODY_FONT
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["C"].width = 60
        return ws

    team_rows = [r for r in rows if r["team"] == team]

    hours_col = _detail_range(COL_HOURS, n_rows)
    team_col = _detail_range(COL_TEAM, n_rows)
    eng_col = _detail_range(COL_ENG, n_rows)
    epic_col = _detail_range(COL_EPIC_KEY, n_rows)

    _write_header(ws, 4, ["Engineer", "Epic key", "Epic", "Hours"], [26, 12, 40, 12])

    pairs, seen = [], set()
    for r in team_rows:
        key = (r["engineer"], r["epic_key"], r["epic_label"])
        if key not in seen:
            seen.add(key)
            pairs.append(key)

    row = 5
    first = row
    for engineer, epic_key, epic_label in pairs:
        ws.cell(row=row, column=1, value=engineer).font = BODY_FONT
        ws.cell(row=row, column=2, value=epic_key).font = BODY_FONT
        ws.cell(row=row, column=3, value=epic_label).font = BODY_FONT
        cell = ws.cell(row=row, column=4)
        cell.value = (f"=SUMIFS({hours_col},{team_col},$A$2,"
                      f"{eng_col},$A{row},{epic_col},$B{row})")
        cell.number_format = HOURS_FMT
        cell.font = BODY_FONT
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = BOX
        row += 1

    total_row = row
    ws.cell(row=total_row, column=3, value="TEAM TOTAL").font = BOLD_FONT
    total = ws.cell(row=total_row, column=4, value=f"=SUM(D{first}:D{row - 1})")
    total.font = BOLD_FONT
    total.number_format = HOURS_FMT
    total.fill = SUBTOTAL_FILL
    for col in range(1, 5):
        ws.cell(row=total_row, column=col).border = BOX

    # Second block: per-engineer capacity and share.
    block = total_row + 3
    ws.cell(row=block, column=1, value="Per engineer").font = SECTION_FONT
    _write_header(
        ws, block + 1,
        ["Engineer", "WBSO hours", "Day equivalent", "Active days",
         "Capacity hours", "WBSO share", "Peak tickets/day"],
        None, freeze=False,
    )
    for col, width in zip("EFG", (15, 13, 18)):
        ws.column_dimensions[col].width = max(ws.column_dimensions[col].width or 0, width)

    day_counts = res.get("day_counts") or {}
    busiest = res.get("busiest") or {}
    engineers, seen_eng = [], set()
    for r in team_rows:
        if r["engineer"] not in seen_eng:
            seen_eng.add(r["engineer"])
            engineers.append(r["engineer"])

    row = block + 2
    eng_first = row
    for engineer in engineers:
        ws.cell(row=row, column=1, value=engineer).font = BODY_FONT
        hours = ws.cell(row=row, column=2)
        hours.value = f"=SUMIFS({hours_col},{team_col},$A$2,{eng_col},$A{row})"
        hours.number_format = HOURS_FMT
        hours.font = BODY_FONT

        day_equiv = ws.cell(row=row, column=3, value=f"=IFERROR($B{row}/{HOURS_PER_DAY_REF},0)")
        day_equiv.number_format = HOURS_FMT
        day_equiv.font = BODY_FONT

        active = ws.cell(row=row, column=4, value=day_counts.get(engineer, 0))
        active.font = INPUT_FONT

        capacity = ws.cell(row=row, column=5, value=f"=$D{row}*{HOURS_PER_DAY_REF}")
        capacity.number_format = HOURS_FMT
        capacity.font = BODY_FONT

        share = ws.cell(row=row, column=6, value=f"=IFERROR($B{row}/$E{row},0)")
        share.number_format = PCT_FMT
        share.font = BODY_FONT

        peak, peak_day = busiest.get(engineer, (0, None))
        ws.cell(row=row, column=7,
                value=f"{peak} on {peak_day:%d %b}" if peak_day else "-").font = INPUT_FONT
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = BOX
        row += 1

    ws.cell(row=row, column=1, value="TOTAL").font = BOLD_FONT
    for col, letter in ((2, "B"), (3, "C"), (4, "D"), (5, "E")):
        cell = ws.cell(row=row, column=col, value=f"=SUM({letter}{eng_first}:{letter}{row - 1})")
        cell.font = BOLD_FONT
        cell.number_format = HOURS_FMT
        cell.fill = SUBTOTAL_FILL
    share = ws.cell(row=row, column=6, value=f"=IFERROR($B{row}/$E{row},0)")
    share.font = BOLD_FONT
    share.number_format = PCT_FMT
    share.fill = SUBTOTAL_FILL
    return ws


def _sheet_detail(wb, rows):
    ws = wb.create_sheet("Detail")
    ws["A1"] = "Per-ticket WBSO hours across every team (input data from Jira)"
    ws["A1"].font = TITLE_FONT
    _write_header(
        ws, DETAIL_FIRST_ROW - 1,
        ["Team", "Engineer", "Epic key", "Epic", "Ticket", "Ticket summary", "Hours"],
        [26, 24, 12, 34, 14, 46, 10],
    )

    row = DETAIL_FIRST_ROW
    for r in rows:
        ws.cell(row=row, column=COL_TEAM, value=r["team"]).font = BODY_FONT
        ws.cell(row=row, column=COL_ENG, value=r["engineer"]).font = BODY_FONT
        ws.cell(row=row, column=COL_EPIC_KEY, value=r["epic_key"]).font = BODY_FONT
        ws.cell(row=row, column=COL_EPIC, value=r["epic_label"]).font = BODY_FONT
        ws.cell(row=row, column=COL_TICKET, value=r["ticket"]).font = BODY_FONT
        ws.cell(row=row, column=COL_SUMMARY, value=r["summary"]).font = BODY_FONT
        hours = ws.cell(row=row, column=COL_HOURS, value=round(r["hours"], 4))
        hours.font = INPUT_FONT
        hours.number_format = HOURS_FMT
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = BOX
        row += 1

    last = row - 1
    if rows:
        total_row = row + 1
        ws.cell(row=total_row, column=COL_SUMMARY, value="TOTAL").font = BOLD_FONT
        total = ws.cell(row=total_row, column=COL_HOURS,
                        value=f"=SUM(G{DETAIL_FIRST_ROW}:G{last})")
        total.font = BOLD_FONT
        total.number_format = HOURS_FMT
        total.fill = SUBTOTAL_FILL
        ws.auto_filter.ref = f"A{DETAIL_FIRST_ROW - 1}:G{last}"
    else:
        ws.cell(row=DETAIL_FIRST_ROW, column=1,
                value="No WBSO hours found for any team in the reporting month.").font = BODY_FONT
    return ws


def _sheet_diagnostics(wb, results):
    ws = wb.create_sheet("Diagnostics")
    ws["A1"] = "Mapping and scope issues found during this run"
    ws["A1"].font = TITLE_FONT
    _write_header(ws, 3, ["Team", "Epic key", "Issue", "Example ticket", "Action"],
                  [26, 14, 42, 16, 52])

    row = 4
    for res in results:
        diag = res.get("diagnostics") or {}
        for epic, ticket in sorted((diag.get("unmapped_epics") or {}).items()):
            for col, value in enumerate([
                res["team"], epic, "Absent from the mapping CSV entirely", ticket,
                "Treated as non-WBSO. Add a row to the CSV if it should count.",
            ], 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = BODY_FONT
                cell.border = BOX
            row += 1
        for epic, ticket in sorted((diag.get("wbso_other_team") or {}).items()):
            for col, value in enumerate([
                res["team"], epic, "Flagged WBSO, but under another team's rows", ticket,
                "Excluded. Re-run with --wbso-all-teams to include it.",
            ], 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = BODY_FONT
                cell.border = BOX
            row += 1
        if res.get("error"):
            for col, value in enumerate([
                res["team"], "", "Team could not be processed", "", res["error"],
            ], 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = BODY_FONT
                cell.border = BOX
            row += 1

    if row == 4:
        ws.cell(row=4, column=1, value="No issues found.").font = BODY_FONT
    return ws


def write_workbook(path, results, meta):
    """
    Build the workbook from a list of per-team result dicts and return
    (path, detail_row_count, {team: sheet_name}).

    Each result: {team, project, month, totals, summaries, day_counts,
                  busiest, diagnostics, wbso_epic_count, status, error}
    """
    wb = Workbook()
    wb.remove(wb.active)

    rows = _flatten(results)
    n_rows = len(rows)

    used = {"assumptions", "summary", "detail", "diagnostics"}
    sheet_names = {res["team"]: safe_sheet_name(res["team"], used) for res in results}

    _sheet_assumptions(wb, meta)
    _sheet_summary(wb, results, n_rows, sheet_names)
    for res in results:
        _sheet_team(wb, res, rows, n_rows, sheet_names[res["team"]])
    _sheet_detail(wb, rows)
    _sheet_diagnostics(wb, results)

    wb.save(path)
    return path, n_rows, sheet_names
